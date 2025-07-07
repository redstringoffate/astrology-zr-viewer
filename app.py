import streamlit as st
import pandas as pd
import swisseph as swe
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Zodiac settings
zr_periods = {
    "Aries":(15,"fire"),"Taurus":(8,"earth"),"Gemini":(20,"air"),"Cancer":(25,"water"),
    "Leo":(19,"fire"),"Virgo":(20,"earth"),"Libra":(8,"air"),"Scorpio":(15,"water"),
    "Sagittarius":(12,"fire"),"Capricorn":(30,"earth"),"Aquarius":(30,"air"),"Pisces":(12,"water")
}
signs = list(zr_periods.keys())
element_colors = {"fire":"FFCCCC","earth":"FFFF99","air":"F2F2F2","water":"CCF0FF"}
sign_glyphs = {"Aries":"♈","Taurus":"♉","Gemini":"♊","Cancer":"♋","Leo":"♌","Virgo":"♍",
               "Libra":"♎","Scorpio":"♏","Sagittarius":"♐","Capricorn":"♑","Aquarius":"♒","Pisces":"♓"}
planet_periods = {"Sun":10,"Venus":8,"Mercury":13,"Moon":9,"Saturn":11,"Jupiter":12,"Mars":7,"Node":3}
planet_glyphs = {"Saturn":"♄","Jupiter":"♃","Mars":"♂","Sun":"☉","Venus":"♀","Mercury":"☿","Moon":"☽","Node":"☊"}
planet_colors = {"♄":"FFFF00","♃":"00B050","♂":"FF0000","☉":"FF99CC","♀":"BFBFBF","☿":"00B0F0","☽":"FFCCFF"}

def zodiac_sign(lon):
    return signs[int(lon//30)]

def lot_position(asc,sun,moon,is_spirit=True):
    return (asc+(sun-moon if is_spirit else moon-sun))%360

def compute_zr(jd_birth,lot_lon,years=100):
    seq,t=[],jd_birth
    lot_sign=zodiac_sign(lot_lon)
    lot_deg=lot_lon%30
    signs_seq=signs[signs.index(lot_sign):]+signs[:signs.index(lot_sign)]
    l1_total,l1_remain=zr_periods[lot_sign][0],zr_periods[lot_sign][0]*(30-lot_deg)/30
    l1_list=[(lot_sign,l1_remain)]
    for s in signs_seq[1:]:
        l1_list.append((s,zr_periods[s][0]))
        if sum(y for _,y in l1_list)>=years: break
    for l1_sign,l1_years in l1_list:
        t1=t+l1_years*365.25
        l2_unit=l1_years/12
        for i,l2_sign in enumerate(signs):
            l2_start=t+i*l2_unit*365.25
            l3_unit=l2_unit/12
            for j,l3_sign in enumerate(signs):
                l3_start=l2_start+j*l3_unit*365.25
                seq.append({"jd":l3_start,"L1":l1_sign,"L2":l2_sign,"L3":l3_sign})
        t=t1
    return seq

def compute_firdaria(jd_birth,is_day,years=100):
    main_seq=(
        ["Sun","Venus","Mercury","Moon","Saturn","Jupiter","Mars","Node"] if is_day
        else ["Moon","Saturn","Jupiter","Mars","Sun","Venus","Mercury","Node"])
    sub_seq = [p for p in main_seq if p != "Node"]
    total_period=73
    seq_main,seq_sub=[],[]
    t=jd_birth
    for m in main_seq:
        mp=planet_periods[m]
        t1=t+mp*365.25
        seq_main.append({"jd":t,"Main":planet_glyphs[m]})
        for s in sub_seq:
            sub_duration=mp*(planet_periods[s]/total_period)*365.25
            seq_sub.append({"jd":t,"Sub":planet_glyphs[s]})
            t+=sub_duration
        t=t1
    return seq_main,seq_sub

def generate_excel(df):
    buffer = BytesIO()
    df.to_excel(buffer, index=False)
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for col in [3,4,5,6,7,8,11,12,13,14]:
            sign=row[col-1].value
            elem=None
            for k,v in sign_glyphs.items():
                if v==sign: elem=zr_periods[k][1]; break
            color=element_colors[elem] if elem else "FFFFFF"
            row[col-1].fill=PatternFill(start_color=color,end_color=color,fill_type="solid")
        for col in [9,10]:
            planet=row[col-1].value
            color=planet_colors.get(planet,"FFFFFF")
            row[col-1].fill=PatternFill(start_color=color,end_color=color,fill_type="solid")
    out = BytesIO()
    wb.save(out)
    return out

st.title("\U0001F52E Zodiacal Releasing + Firdaria + Profections Viewer")

col1, col2 = st.columns(2)
y = col1.number_input("Birth Year", 1900, 2100, 1992)
m = col1.number_input("Month", 1, 12, 6)
d = col1.number_input("Day", 1, 31, 1)
H = col2.number_input("Hour", 0, 23, 9)
mi = col2.number_input("Minute", 0, 59, 30)
lat = col1.number_input("Latitude (e.g. 37.57)", -90.0, 90.0, 37.5665)
lon = col1.number_input("Longitude (e.g. 126.98)", -180.0, 180.0, 126.9780)
tz = col2.number_input("Timezone (e.g. 9)", -12.0, 14.0, 9.0)

if st.button("Generate"):
    local_dt=datetime(int(y),int(m),int(d),int(H),int(mi))
    utc_dt=local_dt-timedelta(hours=tz)
    jd_birth=swe.julday(utc_dt.year,utc_dt.month,utc_dt.day,utc_dt.hour+utc_dt.minute/60)
    swe.set_topo(lon,lat,0)
    asc=swe.houses(jd_birth,lat,lon)[0][0]
    sun_pos=swe.calc_ut(jd_birth,swe.SUN,swe.FLG_SWIEPH|swe.FLG_EQUATORIAL)[0]
    is_day=sun_pos[1]>0
    sun=swe.calc_ut(jd_birth,swe.SUN)[0][0]
    moon=swe.calc_ut(jd_birth,swe.MOON)[0][0]
    lot_fort=lot_position(asc,sun,moon,is_spirit=False)
    lot_spir=lot_position(asc,sun,moon,is_spirit=True)
    fort_seq=compute_zr(jd_birth,lot_fort)
    spir_seq=compute_zr(jd_birth,lot_spir)
    fird_main,fird_sub=compute_firdaria(jd_birth,is_day)
    dates=sorted(set([f["jd"]for f in fort_seq]+[s["jd"]for s in spir_seq]+
                     [m["jd"]for m in fird_main]+[s["jd"]for s in fird_sub]))
    fi,si,mi,si2=0,0,0,0
    rows=[]
    for jd in dates:
        while fi+1<len(fort_seq) and fort_seq[fi+1]["jd"]<=jd: fi+=1
        while si+1<len(spir_seq) and spir_seq[si+1]["jd"]<=jd: si+=1
        while mi+1<len(fird_main) and fird_main[mi+1]["jd"]<=jd: mi+=1
        while si2+1<len(fird_sub) and fird_sub[si2+1]["jd"]<=jd: si2+=1
        date_dt=datetime(*swe.revjul(jd)[:3])
        age=int((jd-jd_birth)/365.25)
        prof_sign_idx=(int(asc//30)+(age%12))%12
        prof_sign=signs[prof_sign_idx]
        prof_sign_glyph=sign_glyphs[prof_sign]
        uranus_lon=swe.calc_ut(jd,swe.URANUS)[0][0]
        neptune_lon=swe.calc_ut(jd,swe.NEPTUNE)[0][0]
        pluto_lon=swe.calc_ut(jd,swe.PLUTO)[0][0]
        uranus_sign=sign_glyphs[zodiac_sign(uranus_lon)]
        neptune_sign=sign_glyphs[zodiac_sign(neptune_lon)]
        pluto_sign=sign_glyphs[zodiac_sign(pluto_lon)]
        rows.append({"Date":date_dt.strftime("%Y-%m-%d"),"Age":age,
                     "Fortune_L1":sign_glyphs[fort_seq[fi]["L1"]],"Fortune_L2":sign_glyphs[fort_seq[fi]["L2"]],"Fortune_L3":sign_glyphs[fort_seq[fi]["L3"]],
                     "Spirit_L1":sign_glyphs[spir_seq[si]["L1"]],"Spirit_L2":sign_glyphs[spir_seq[si]["L2"]],"Spirit_L3":sign_glyphs[spir_seq[si]["L3"]],
                     "Firdaria_Main":fird_main[mi]["Main"],"Firdaria_Sub":fird_sub[si2]["Sub"],
                     "Profection_Sign":prof_sign_glyph,
                     "Uranus":uranus_sign,"Neptune":neptune_sign,"Pluto":pluto_sign})
    df=pd.DataFrame(rows)
    st.success("Calculation complete. Showing first 10 rows below:")
    st.dataframe(df.head(10))
    excel_data = generate_excel(df)
    st.download_button("📄 Download Excel File", data=excel_data.getvalue(), file_name="ZR_Firdaria_Profections.xlsx")
